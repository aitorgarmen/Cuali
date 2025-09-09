# tabs/frequency_scan.py
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import QLabel
from PyQt5.QtCore import Qt
from typing import Tuple, List, Dict, Any, Optional
import serial

# define max 32-bit int
MAX_32BIT_INT = 2**31 - 1

import numpy as np, itertools, time, pyqtgraph as pg
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, styles
from openpyxl.styles import PatternFill
from serial.tools import list_ports

from device import Device, _pack32
from utils import labeled_spinbox, thin_separator
from style import apply_app_style   # solo si necesitas colores pg aquí

# ------------- ScanWorker (QThread) -----------------------------------------
class ScanWorker(QtCore.QThread):
    frame_ready  = QtCore.pyqtSignal(dict)                 # datos → UI
    progress     = QtCore.pyqtSignal(int)                  # porcentaje 0-100
    param_update = QtCore.pyqtSignal(float, float, float)  # f, g, p actuales
    finished     = QtCore.pyqtSignal(list)                 # rows (list[dict])
    error        = QtCore.pyqtSignal(str)

    def __init__(self, device: 'Device',
                 ranges: Tuple[np.ndarray, np.ndarray, np.ndarray],
                 const_params: Dict[str, Any],
                 algo_params: Tuple[np.uint32, np.uint32],
                 bolt_num: float):
        super().__init__()
        self._device = device
        self._f_list, self._g_list, self._p_list = ranges
        self._const = const_params
        self._algo  = algo_params                   # (long_algo, long_umbral)
        self._bolt  = bolt_num
        self._stop  = False
        self._rows: list = []

    # .........................................................................
    def stop(self):
        self._stop = True

    # .........................................................................
    def run(self):
        try:
            self._send_constant_config()
            total = len(self._f_list) * len(self._g_list) * len(self._p_list)
            it    = 0

            for f, g, p in itertools.product(self._f_list,
                                              self._g_list,
                                              self._p_list):
                if self._stop:
                    break
                # Aviso UI de los parámetros que vamos a probar
                self.param_update.emit(f, g, p)
                self._send_variable_config(f, g, p)
                frame = self._acquire_frame()
                if not frame:
                    break

                # registro para Excel: incluir temperatura entre tof y force
                self._rows.append([
                    f, g, p, self._bolt,
                    frame["dat3"], frame["dat2"],
                    frame["pico1"], frame["porcentaje_diferencia"],
                    frame["tof"], frame.get("temp"), frame["force"],
                    frame.get("maxcorrx"), frame.get("maxcorry")
                ])

                it += 1
                self.progress.emit(int(it / total * 100))
                self.frame_ready.emit(frame)

                # breve pausa para no saturar la CPU (≈ 5-10 ms)
                self.msleep(5)

            self.finished.emit(self._rows)

        except Exception as e:
            self.error.emit(str(e))
        finally:
            self._device.ser.close()

    # --------------------- helpers internos ----------------------------------
    def _send_constant_config(self):
        d = self._device
        p = self._const
        long_algo, long_umbral = self._algo

        d.modo_standby()
        d.modo_configure()

        # ----- parámetros que el FW borra al salir de modo_single -----
        d.enviar(_pack32(20.0),       "10")
        d.enviar(_pack32(p["diftemp"]),    "11")
        d.enviar(_pack32(p["tof"]),        "12")
        # constantes de correcciones y threshold
        d.enviar(_pack32(p["xi"]),         "18")
        d.enviar(_pack32(p["alpha"]),      "19")
        d.enviar(_pack32(p["short_corr"]), "1A")
        d.enviar(_pack32(p["long_corr"]),  "1B")
        d.enviar(_pack32(p["short_temp"]), "1C")
        d.enviar(_pack32(p["long_temp"]),  "1D")
        d.enviar(_pack32(long_algo),       "2C")
        d.enviar(_pack32(long_umbral),     "2D")

        d.modo_save()
        d.modo_standby()
        d.modo_single()

    # .........................................................................
    def _send_variable_config(self, freq, gain, pulse):
        d = self._device
        d.modo_configure()
        d.enviar(_pack32(freq),  "14")
        d.enviar(_pack32(pulse), "15")
        d.enviar(_pack32(pulse), "16")
        d.enviar(_pack32(gain),  "17")
        d.modo_save()
        d.modo_standby()
        d.modo_single()

    # .........................................................................
    def _acquire_frame(self):
        d = self._device
        try:
            d.enviar_temp()
            d.start_measure()
            return d.lectura()
        except (serial.SerialException, PermissionError) as e:
            self.error.emit(f"Puerto perdido: {e}. Reintentando...")
            try:
                d.ser.close()
                time.sleep(0.2)
                d.ser.open()
            except Exception:
                self._stop = True          # aborta el hilo si no se pudo recuperar
            return {}


# ------------- FrequencyScanTab ------------------------------------
class FrequencyScanTab(QtWidgets.QWidget):
    def __init__(
        self,
        com_selector: QtWidgets.QComboBox | None = None,
        *,
        db: Optional[Any] = None,
        batch_id: Optional[str] = None,
    ):
        super().__init__()
        # Store COM selector and database info for later use
        self.com_selector = com_selector
        self.db = db
        self.batch_id = batch_id
        self._build_ui()
        self._worker: ScanWorker | None = None
        self._device: 'Device' | None = None
        self._rows: list | None = None
        # self.com_selector: QtWidgets.QComboBox | None = None

    # .........................................................................
    def set_com_selector(self, combo):
        self.com_selector = combo

    # =======================  UI  ============================================
    def _build_ui(self):
        # ------------------------------------------------------------------ #
        #  PANEL IZQUIERDO - Parámetros de escaneado
        # ------------------------------------------------------------------ #
        param_layout = QtWidgets.QFormLayout()
        param_layout.setLabelAlignment(Qt.AlignLeft)
        param_layout.setFormAlignment(Qt.AlignTop)
        param_layout.setVerticalSpacing(14)

        # Encabezados "Inicio / Final / Step"
        header = QtWidgets.QWidget()
        h_head = QtWidgets.QHBoxLayout(header); h_head.setContentsMargins(0, 0, 0, 0)
        h_head.addSpacing(100)
        for txt in ("Inicio", "Final", "Step"):
            lab = QtWidgets.QLabel(txt); lab.setAlignment(Qt.AlignCenter)
            h_head.addWidget(lab, 1)
        param_layout.addRow(header)

        # ---- Spinners de rango (freq / gain / pulse) ---------------------
        self.freq_start = labeled_spinbox(2, 100, step=1, decimals=0)
        self.freq_end   = labeled_spinbox(2, 100, step=1, decimals=0)
        self.freq_step  = labeled_spinbox(1, 10,  step=1, decimals=0)

        self.gain_start = labeled_spinbox(5, 80,  step=1, decimals=0)
        self.gain_end   = labeled_spinbox(5, 80,  step=1, decimals=0)
        self.gain_step  = labeled_spinbox(1, 10,  step=1, decimals=0)

        self.pulse_start= labeled_spinbox(1, 16, step=1, decimals=0)
        self.pulse_end  = labeled_spinbox(1, 16, step=1, decimals=0)
        # self.pulse_step = labeled_spinbox(1,  4, step=1, decimals=0)
        self.pulse_step = labeled_spinbox(1,  6, step=1, decimals=0)

        # self.freq_start.setValue(30); self.freq_end.setValue(50); self.freq_step.setValue(2)
        # self.gain_start.setValue(10); self.gain_end.setValue(30); self.gain_step.setValue(5)
        # self.pulse_start.setValue(1); self.pulse_end.setValue(4); self.pulse_step.setValue(1)

        # self.freq_start.setValue(24); self.freq_end.setValue(36); self.freq_step.setValue(1)
        # self.gain_start.setValue(25); self.gain_end.setValue(25); self.gain_step.setValue(5)
        # self.pulse_start.setValue(10); self.pulse_end.setValue(16); self.pulse_step.setValue(1)

        self.freq_start.setValue(24); self.freq_end.setValue(36); self.freq_step.setValue(2)
        self.gain_start.setValue(25); self.gain_end.setValue(25); self.gain_step.setValue(5)
        self.pulse_start.setValue(10); self.pulse_end.setValue(16); self.pulse_step.setValue(2)

        def _row(lbl, widgets):
            h = QtWidgets.QHBoxLayout()
            [h.addWidget(w) for w in widgets]
            param_layout.addRow(lbl, h)

        _row("Frequency", (self.freq_start, self.freq_end, self.freq_step))
        _row("Gain",       (self.gain_start,  self.gain_end,  self.gain_step))
        _row("Pulse",      (self.pulse_start, self.pulse_end, self.pulse_step))
        param_layout.addRow(thin_separator())

        # ---- Valores únicos ------------------------------------------------
        self.temperature = labeled_spinbox(-22, 50, step=0.1, decimals=1)
        # ToF mode: manual or calculate
        self.tof_mode    = QtWidgets.QComboBox()
        self.tof_mode.addItems(["Manual", "Calculate"])
        self.tof_mode.setCurrentIndex(1)  # por defecto, calcular ToF
        self.tof         = labeled_spinbox(0, MAX_32BIT_INT, step=0.1, decimals=1)
        # Calculation fields and labels (hidden by default)
        self.bolt_length_label = QLabel("Bolt Length (mm)")
        self.bolt_length = labeled_spinbox(0, 10000, step=1, decimals=1)
        # self.bolt_length.setValue(720.0)
        self.bolt_length.setValue(714.0)
        self.velocity_label = QLabel("Velocity (m/s)")
        self.velocity    = labeled_spinbox(1000, 10000, step=10, decimals=1)
        self.velocity.setValue(5900.0)
        self.bolt_length_label.hide(); self.bolt_length.hide()
        self.velocity_label.hide();    self.velocity.hide()
        self.bolt        = labeled_spinbox(0, MAX_32BIT_INT, step=1,   decimals=0)

        self.temperature.setValue(20.0)
        self.tof.setValue(244500.0)
        self.bolt.setValue(0)

        param_layout.addRow("Temperature", self.temperature)
        param_layout.addRow("ToF mode",     self.tof_mode)
        param_layout.addRow("ToF",           self.tof)
        param_layout.addRow(self.bolt_length_label, self.bolt_length)
        param_layout.addRow(self.velocity_label,    self.velocity)
        param_layout.addRow("Bolt Num",     self.bolt)

        # ---- Algoritmo + Threshold ----------------------------------------
        self.algo_selector = QtWidgets.QComboBox()
        self.algo_selector.addItems(["Absolute maximum", "First maximum", "Second maximum"])
        param_layout.addRow("Algorithm", self.algo_selector)

        self.threshold_label = QtWidgets.QLabel("Threshold")
        self.threshold_spin  = QtWidgets.QDoubleSpinBox()
        self.threshold_spin.setRange(0.0, 1e6)
        self.threshold_spin.setDecimals(2)
        self.threshold_spin.setSingleStep(1.0)
        self.threshold_label.setVisible(False)
        self.threshold_spin.setVisible(False)
        param_layout.addRow(self.threshold_label, self.threshold_spin)

        # ---- Botones PLAY / STOP ------------------------------------------
        self.btn_play = QtWidgets.QPushButton("PLAY", objectName="Play")
        self.btn_stop = QtWidgets.QPushButton("STOP", objectName="Stop")
        bb = QtWidgets.QHBoxLayout(); bb.addWidget(self.btn_play); bb.addWidget(self.btn_stop)
        param_layout.addRow(bb)

        # ---- Progress Bar de progreso -------------------------------------
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("%p%")
        self.progress_bar.setTextVisible(True)
        param_layout.addRow("Progress", self.progress_bar)
        # ---- Parámetros actuales ------------------------------------------
        self.param_label = QLabel("Freq: -   Gain: -   Pulse: -")
        param_layout.addRow("Params", self.param_label)

        # ---- Contenedor izquierdo -----------------------------------------
        left = QtWidgets.QFrame(); left.setLayout(param_layout)

        # ------------------------------------------------------------------ #
        #  PANEL DERECHO - Fichero, métricas y gráfica
        # ------------------------------------------------------------------ #
        right = QtWidgets.QVBoxLayout(); right.setSpacing(12)

        # ---- Selector de fichero ------------------------------------------
        fbox = QtWidgets.QHBoxLayout()
        self.file_edit = QtWidgets.QLineEdit()
        self.format_selector = QtWidgets.QComboBox()
        self.format_selector.addItems(["CSV (.csv)", "Excel (.xlsx)"])
        self.file_btn  = QtWidgets.QPushButton("File", objectName="File")
        fbox.addWidget(self.file_edit)
        fbox.addWidget(self.format_selector)
        fbox.addWidget(self.file_btn)
        right.addLayout(fbox)

        # ---- Métricas ------------------------------------------------------
        mbox = QtWidgets.QHBoxLayout(); mbox.setSpacing(20)
        mbox.addWidget(QtWidgets.QLabel("Amp. max:")); self.amp_max = QtWidgets.QLineEdit(); mbox.addWidget(self.amp_max)
        mbox.addWidget(QtWidgets.QLabel("Pct. dif:")); self.pct_dif = QtWidgets.QLineEdit(); mbox.addWidget(self.pct_dif)
        mbox.addWidget(QtWidgets.QLabel("ToF:"));      self.tof_edit = QtWidgets.QLineEdit(); mbox.addWidget(self.tof_edit)
        right.addLayout(mbox)
        # ---- Data Type selector for plot ----
        h_data = QtWidgets.QHBoxLayout()
        h_data.addWidget(QtWidgets.QLabel("Data Type:"))
        self.data_cb = QtWidgets.QComboBox()
        self.data_cb.addItems(["dat2","dat3"])
        # default to dat3
        self.data_cb.setCurrentIndex(1)
        h_data.addWidget(self.data_cb)
        right.addLayout(h_data)

        # ---- Gráfica -------------------------------------------------------
        self.plot = pg.PlotWidget(title="Correlated signal")
        self.plot.setMinimumHeight(350)
        self.plot.showGrid(x=True, y=True, alpha=0.2)
        self.plot.setLabel('bottom', 'Samples'); self.plot.setLabel('left', 'Amplitude')
        self.curve = self.plot.plot([], [], pen=pg.mkPen('#00c853', width=2))
        self.marker = pg.ScatterPlotItem(size=8,
                                         pen=pg.mkPen('#ff5722', width=2),
                                         brush=pg.mkBrush('#ff5722'),
                                         symbol='o')
        self.plot.addItem(self.marker); self.marker.setVisible(False)
        self.threshold_line = pg.InfiniteLine(angle=0,
                                              pen=pg.mkPen('#FFEA00', width=1, style=Qt.DashLine),
                                              movable=False)
        self.plot.addItem(self.threshold_line); self.threshold_line.setVisible(False)
        # Disable mouse zoom/pan to avoid unintended zooms
        try:
            _vb = self.plot.getPlotItem().getViewBox()
            _vb.setMouseEnabled(False, False)
            _vb.setMenuEnabled(False)
            # swallow common zoom interactions
            self.plot.wheelEvent = lambda ev: None
            _vb.wheelEvent = lambda ev: None
            _vb.mouseDoubleClickEvent = lambda ev: None
            _vb.mouseDragEvent = lambda ev: None
        except Exception:
            pass
        right.addWidget(self.plot, 1)

        # ---- Opciones de visualización ------------------------------------
        h_mode = QtWidgets.QHBoxLayout()
        h_mode.addWidget(QtWidgets.QLabel("Axes mode:"))
        self.axis_mode = QtWidgets.QComboBox(); self.axis_mode.addItems(["Dynamic", "Static"])
        h_mode.addWidget(self.axis_mode)
        right.addLayout(h_mode)

        h_ylim = QtWidgets.QHBoxLayout(); h_ylim.addWidget(QtWidgets.QLabel("Y limit:"))
        self.ylim_spin = QtWidgets.QDoubleSpinBox()
        self.ylim_spin.setRange(0.0, 1e6); self.ylim_spin.setDecimals(2); self.ylim_spin.setSingleStep(1.0)
        self.ylim_spin.setAlignment(Qt.AlignCenter); self.ylim_spin.setValue(100000.0)
        h_ylim.addWidget(self.ylim_spin)
        self.ylim_container = QtWidgets.QWidget(); self.ylim_container.setLayout(h_ylim); self.ylim_container.setVisible(False)
        right.addWidget(self.ylim_container)

        # ---- Checkbox mostrar gráfica -------------------------------------
        self.chk_show = QtWidgets.QCheckBox("Show graph")
        # show graph by default
        self.chk_show.setChecked(True)
        self.chk_show.stateChanged.connect(lambda st: self.curve.setVisible(bool(st)))
        right.addWidget(self.chk_show, alignment=Qt.AlignRight)

        # ------------------------------------------------------------------ #
        #  ENSAMBLA TRIPLAYOUT PRINCIPAL
        # ------------------------------------------------------------------ #
        # Usar layout explícito y asignarlo al widget
        main = QtWidgets.QHBoxLayout()
        main.addWidget(left)
        main.addLayout(right, 1)
        self.setLayout(main)

        # ------------------------------------------------------------------ #
        #  CONEXIONES
        # ------------------------------------------------------------------ #
        self.btn_play.clicked.connect(self._on_play)
        self.btn_stop.clicked.connect(self._on_stop)
        self.file_btn.clicked.connect(self._choose_file)

        # Algoritmo → mostrar threshold
        self.algo_selector.currentIndexChanged.connect(
            lambda idx: (self.threshold_label.setVisible(idx in (1, 2)),
                         self.threshold_spin.setVisible(idx in (1, 2))))

        # Axes mode callbacks
        self.axis_mode.currentIndexChanged.connect(
            lambda idx: (
                self.ylim_container.setVisible(idx == 1),
                self.ylim_spin.setEnabled(idx == 1),
                self.plot.getPlotItem().enableAutoRange(idx == 0, idx == 0),
                self.plot.getPlotItem().setYRange(-self.ylim_spin.value(),
                                                  self.ylim_spin.value(),
                                                  padding=0) if idx == 1 else None
            ))
        self.ylim_spin.valueChanged.connect(
            lambda v: (self.plot.getPlotItem()
                       .setYRange(-v, v, padding=0) if self.axis_mode.currentIndex() == 1 else None))

        # ToF mode logic: show/hide fields and calculate round-trip ToF in µs
        def update_tof():
            # calculate ToF = 2*length(m)/vel(m/s) in µs
            length_mm = float(self.bolt_length.value())
            vel_m_s   = float(self.velocity.value())
            tof_us    = (2*length_mm) / vel_m_s * 1e6
            self.tof.setValue(tof_us)
        def toggle_tof_mode():
            manual = self.tof_mode.currentText() == "Manual"
            self.tof.setEnabled(manual)
            self.bolt_length_label.setVisible(not manual)
            self.bolt_length.setVisible(not manual)
            self.velocity_label.setVisible(not manual)
            self.velocity.setVisible(not manual)
            if not manual:
                update_tof()
        self.tof_mode.currentIndexChanged.connect(lambda _: toggle_tof_mode())
        self.bolt_length.valueChanged.connect(lambda _: update_tof() if self.tof_mode.currentText()=="Calculate" else None)
        self.velocity.valueChanged.connect(lambda _: update_tof() if self.tof_mode.currentText()=="Calculate" else None)
        # initial state
        toggle_tof_mode()

        # Estado inicial de botones
        self.btn_stop.setEnabled(False)


    # =======================  PLAY / STOP  ==================================
    def _on_play(self):
        self._manual_stop = False
        # --- Cierra el puerto previo si aún estaba abierto ---
        if getattr(self, "_device", None) and self._device.ser.is_open:
            try:
                self._device.ser.close()
            except Exception:
                pass
            self._device = None

        if not self.com_selector or not self.com_selector.currentText():
            QtWidgets.QMessageBox.warning(self, "COM", "Selecciona un puerto válido.")
            return
        try:
            self._device = Device(self.com_selector.currentText(), baudrate=115200, timeout=1)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "UART", str(e)); return
        
        # ── 1. el usuario debe elegir un .xlsx ──────────────────────────
        file_path = self.file_edit.text().strip()
        if not file_path:
            QtWidgets.QMessageBox.warning(
                self, "Sin fichero",
                "Selecciona el fichero .xlsx donde guardar los datos antes de PLAY."
            )
            return


        self.btn_play.setEnabled(False); self.btn_stop.setEnabled(True)
        # Reset progress bar
        self.progress_bar.setValue(0)

        # RANGOS
        f_list = np.arange(self.freq_start.value(),
                           self.freq_end.value() + 0.1,
                           self.freq_step.value(), dtype=np.float32)
        g_list = np.arange(self.gain_start.value(),
                           self.gain_end.value() + 0.1,
                           self.gain_step.value(), dtype=np.float32)
        p_list = np.arange(self.pulse_start.value(),
                           self.pulse_end.value() + 0.1,
                           self.pulse_step.value(), dtype=np.float32)

        # CONST + ALGO
        diftemp, long_corr, xi, alpha = -103.0, 10, 0.0, 0.0
        if self.db and self.batch_id:
            try:
                diftemp, long_corr, xi, alpha = self.db.get_device_params(self.batch_id)
            except Exception:
                pass
        # Resolve window params from DB; fallback to current defaults used here
        scw = stw = ltw = None
        if self.db and self.batch_id:
            try:
                attrs = (self.db.get_batch(self.batch_id) or {}).get("attrs", {})
                def _to_int(v):
                    try:
                        return int(float(v))
                    except Exception:
                        return None
                scw = _to_int(attrs.get("short_correlation_window"))
                stw = _to_int(attrs.get("short_temporal_window"))
                ltw = _to_int(attrs.get("long_temporal_window"))
            except Exception:
                scw = stw = ltw = None
        const = dict(
            temp=float(self.temperature.value()),
            diftemp=np.float32(diftemp),
            tof=float(self.tof.value()),
            xi=np.float32(xi),
            alpha=np.float32(alpha),
            short_corr=np.uint32(scw if scw is not None else 1000),
            long_corr=np.uint32(long_corr),
            short_temp=np.uint32(stw if stw is not None else 1000),
            long_temp=np.uint32(ltw if ltw is not None else 990),
        )
        algo_idx = self.algo_selector.currentIndex()
        if algo_idx == 0:
            long_algo, long_umbral = np.uint32(0), np.uint32(0)
        else:
            long_algo   = np.uint32(algo_idx)
            long_umbral = np.uint32(self.threshold_spin.value())

        # Hilo
        self._worker = ScanWorker(
            self._device,
            ranges=(f_list, g_list, p_list),
            const_params=const,
            algo_params=(long_algo, long_umbral),
            bolt_num=float(self.bolt.value())
        )
        self._worker.frame_ready.connect(self._update_graph,  QtCore.Qt.QueuedConnection)
        self._worker.progress.connect(self._update_progress, QtCore.Qt.QueuedConnection)
        self._worker.param_update.connect(self._update_params, QtCore.Qt.QueuedConnection)
        self._worker.finished.connect(self._on_finished,      QtCore.Qt.QueuedConnection)
        self._worker.error.connect(self._on_error,            QtCore.Qt.QueuedConnection)
        self._worker.start()

    def _update_params(self, freq: float, gain: float, pulse: float):
        """Muestra en UI los valores de freq, gain y pulse en curso."""
        self.param_label.setText(f"Freq: {freq:.0f}   Gain: {gain:.0f}   Pulse: {pulse:.0f}")

    # .........................................................................
    def _on_stop(self):
        self._manual_stop = True              # ← marca que es un stop manual
        self.btn_stop.setEnabled(False)

        if self._worker and self._worker.isRunning():
            self._worker.stop()
            self._worker.wait()
            self._worker = None

        # cierra puerto si sigue abierto
        if getattr(self, "_device", None) and self._device.ser.is_open:
            self._device.ser.close()
            self._device = None

        # Reset progress bar
        self.progress_bar.setValue(0)

        self.btn_play.setEnabled(True)



    # =======================  CALLBACKS  =====================================
    def _update_progress(self, pct: int):
        # Update progress bar value (displays percentage)
        self.progress_bar.setValue(pct)

    def _update_graph(self, frame: dict):
        # Métricas
        self.amp_max.setText(str(frame["maxcorry"]))
        self.pct_dif.setText(f"{frame['porcentaje_diferencia']:.2f}")
        self.tof_edit.setText(f"{frame['tof']:.1f}")

        if not self.chk_show.isChecked():
            self.curve.setData([], [])
            return

        # Plot selected data type
        dtype = self.data_cb.currentText()
        data = frame.get(dtype, frame.get("dat3", []))
        self.curve.setData(data, autoDownsample=True)

        # Show threshold line if applicable
        if self.algo_selector.currentIndex() in (1, 2):
            self.threshold_line.setValue(float(self.threshold_spin.value()))
            self.threshold_line.setVisible(True)
        else:
            self.threshold_line.setVisible(False)
        # Show marker only when plotting dat3
        if self.data_cb.currentText() == 'dat3':
            self.marker.setData([frame["maxcorrx"]], [frame["maxcorry"]])
            self.marker.setVisible(True)
        else:
            self.marker.setVisible(False)
        # Emit beep en cada iteración para señal sonora
        QtWidgets.QApplication.beep()
        # Pitido en cada iteración de Frequency Scan
        QtWidgets.QApplication.beep()

    def _on_finished(self, rows: list):
        self.btn_play.setEnabled(True)
        self.btn_stop.setEnabled(False)

        # resetea progresos siempre
        self.progress_bar.setValue(0)

        if self._manual_stop:
            # el usuario canceló → no guardamos nada
            return

        if rows:                              # solo si hay datos completos
            self._save_to_excel(rows)

    def _on_error(self, msg: str):
        QtWidgets.QMessageBox.critical(self, "Error", msg)
        self._on_stop()

    # =======================  FILE & EXCEL  ==================================
    def _choose_file(self):
        # Defino y creo la carpeta "Medidas" un nivel arriba del script
        medidas_dir = Path(__file__).parent.parent / "Medidas"
        medidas_dir.mkdir(parents=True, exist_ok=True)
        # Nombre por defecto dentro de esa carpeta según formato
        fmt = self.format_selector.currentText()
        if fmt.startswith("CSV"):
            default_path = str(medidas_dir / "M64x700.csv")
            filter_str = "CSV files (*.csv)"
        else:
            default_path = str(medidas_dir / "M64x700.xlsx")
            filter_str = "Archivos de Excel (*.xlsx)"
        name, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Guardar como", default_path, filter_str)
        if name and not name.lower().endswith(
            (".xlsx" if filter_str.endswith("*.xlsx)") else ".csv")):
            name += ".xlsx" if filter_str.endswith("*.xlsx)\"") else ".csv"
        self.file_edit.setText(name)

    # FrequencyScanTab._save_to_excel
    def _save_to_excel(self, data_rows: list):
        file_path = self.file_edit.text().strip()
        if not file_path:
            QtWidgets.QMessageBox.warning(self, "Sin fichero",
                                        "Selecciona .xlsx donde guardar los datos.")
            return

        # ------------------ DataFrame --------------------------------------
        rows = []
        # data_rows now includes temperature between tof and force
        for f, g, p, bolt, dat3, dat2, pico1, pct, tof, temp, force, maxx, maxy in data_rows:
            # Build record dict with temp inserted between tof and force
            rec = dict(
                Freq=f, Gain=g, Pulse=p, **{"Bolt Num": bolt},
                pico1=pico1, pct_diff=pct, tof=tof, temp=temp,
                force=force, maxcorrx=maxx, maxcorry=maxy
            )
            rec.update({f"dat3_{i}": v for i, v in enumerate(dat3)})
            rec.update({f"dat2_{i}": v for i, v in enumerate(dat2)})
            rows.append(rec)
        df = pd.DataFrame(rows)
        df['pct_diff'] /= 100.0
        # Si CSV, guarda sin formato y sale
        if file_path.lower().endswith('.csv'):
            # Convertir valores de Freq, Gain, Pulse y Bolt Num a int para CSV
            df[['Freq','Gain','Pulse','Bolt Num']] = df[['Freq','Gain','Pulse','Bolt Num']].astype(int)
            csv_path = Path(file_path)
            csv_path.parent.mkdir(parents=True, exist_ok=True)
            # Append or write header based on file existence
            mode = 'a' if csv_path.exists() and csv_path.stat().st_size > 0 else 'w'
            header = not (mode == 'a')
            df.to_csv(file_path, index=False, mode=mode, header=header)
            QtWidgets.QMessageBox.information(self, "Guardado", f"Datos guardados en CSV:\n{file_path}")
            return

        # ------------------ Writer seguro para Excel -----------------------
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        file_exists = Path(file_path).exists()

        writer_kw = dict(engine='openpyxl', mode=('a' if file_exists else 'w'))
        if file_exists:
            writer_kw['if_sheet_exists'] = 'overlay'

        startrow = 0
        if file_exists:
            wb  = load_workbook(file_path)
            ws  = wb['Datos'] if 'Datos' in wb.sheetnames else wb.create_sheet('Datos')
            startrow = ws.max_row
            wb.close()                     # muy importante: cierra el handle

        with pd.ExcelWriter(file_path, **writer_kw) as wr:
            df.to_excel(wr, index=False, sheet_name='Datos',
                        startrow=startrow, header=not file_exists)

        # ------------------ formato condicional ----------------------------
        # usamos load_workbook y PatternFill ya importados a nivel de módulo
        green_fill = PatternFill(fill_type='solid', fgColor='C6EFCE')
        wb = load_workbook(file_path)
        ws = wb['Datos']
        pct_col   = df.columns.get_loc('pct_diff') + 1
        tof_col   = df.columns.get_loc('tof')      + 1
        force_col = df.columns.get_loc('force')    + 1
        max_row   = ws.max_row
        max_col   = ws.max_column
        start_row = max_row - len(df) + 1

        for row in range(start_row, max_row + 1):
            # formateo de celdas numéricas
            ws.cell(row=row, column=pct_col).number_format   = '0.0%'
            ws.cell(row=row, column=tof_col).number_format   = '0.0'
            ws.cell(row=row, column=force_col).number_format = '0.0'
            # relleno condicional si pct_diff > 20%
            val = ws.cell(row=row, column=pct_col).value or 0
            if val > 0.20:
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = green_fill
        wb.save(file_path)
        
        QtWidgets.QMessageBox.information(self, "Guardado", f"Datos guardados en:\n{file_path}")
