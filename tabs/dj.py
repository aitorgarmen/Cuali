"""
tabs/dj.py

Pestaña DJ: prueba de ciclos 'DJ' en el dispositivo.
"""

from PyQt5 import QtWidgets, QtCore
from PyQt5.QtCore import Qt
import time

from PyQt5.QtWidgets import QLabel, QFileDialog
from typing import Callable, Dict, Any, Optional
import numpy as np
import pandas as pd
import pyqtgraph as pg
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from device import Device, _pack32  # import Device and pack utility
from PyQt5.QtWidgets import QFileDialog, QLabel, QDoubleSpinBox, QComboBox

from utils import thin_separator, labeled_spinbox

# define max 32-bit int
MAX_32BIT_INT = 2**31 - 1

# ---------------------------------------------------------------------------
# 1) Thread worker
# ---------------------------------------------------------------------------
class DeviceWorker(QtCore.QThread):
    """Hilo que gestiona la conversación con el dispositivo."""
    data_ready = QtCore.pyqtSignal(dict)   # frame completo → UI
    error      = QtCore.pyqtSignal(str)    # mensaje de error → UI

    def __init__(self, device: Device, param_getter: Callable[[], Dict[str, Any]],
                 db: Optional[Any] = None, batch_id: Optional[str] = None, parent=None):
        super().__init__(parent)
        self._device = device
        self._param_getter = param_getter
        self._stop = False
        self._cached_params: Dict[str, Any] | None = None

        # Constantes del protocolo
        diftemp, long_corr, xi, alpha = -103.0, 10, 0.0, 0.0
        if db and batch_id:
            try:
                diftemp, long_corr, xi, alpha = db.get_device_params(batch_id)
            except Exception:
                pass
        self._diftemp    = np.float32(diftemp)
        self._xi         = np.float32(xi)
        self._alpha      = np.float32(alpha)
        # Defaults as currently used in DJ
        short_corr_def = 1000
        short_temp_def = 1000
        long_temp_def  = 990
        scw = stw = ltw = None
        if db and batch_id:
            try:
                b = db.get_batch(batch_id) or {"attrs": {}}
                attrs = b.get("attrs", {}) or {}
                def _to_int(v):
                    try:
                        return int(float(v))
                    except Exception:
                        return None
                scw = _to_int(attrs.get("short_correlation_window"))
                stw = _to_int(attrs.get("short_temporal_window"))
                ltw = _to_int(attrs.get("long_temporal_window"))
            except Exception:
                scw = stw = ltw = None
        self._short_corr = np.uint32(scw if scw is not None else short_corr_def)
        self._long_corr  = np.uint32(long_corr)
        self._short_temp = np.uint32(stw if stw is not None else short_temp_def)
        self._long_temp  = np.uint32(ltw if ltw is not None else long_temp_def)

    # ------------------------------------------------ run (bucle principal)
    def run(self):
        try:
            while not self._stop:
                p = self._param_getter()  # snapshot de los valores actuales

                # 1️⃣ Si han cambiado parámetros => envío configuración completa
                if p != self._cached_params:
                    self._send_full_config(p)
                    self._cached_params = p

                # 2️⃣ Disparo de medida y lectura
                frame = self._acquire_frame()
                self.data_ready.emit(frame)       # "fire‑and‑forget", cola Qt

                # 3️⃣ Pequeño respiro para evitar busy‑loop si el equipo es rápido
                self.msleep(10)   # ~10 ms → "búfer" de UI y CPU
        finally:
            # Nos aseguramos de cerrar el puerto aunque el hilo muera por error
            self._device.ser.close()

    # ------------------------------------------------ helpers internos
    def stop(self):
        self._stop = True

    # .................................................................
    def _send_full_config(self, p: Dict[str, Any]):
        """Envía todos los registros sólo si algo ha cambiado."""
        algo_idx = p["algo"]
        if algo_idx == 0:          # Absolute maximum
            long_algo, long_umbral = np.uint32(0), np.uint32(0)
        else:                      # First/Second maximum
            long_algo   = np.uint32(algo_idx)
            long_umbral = np.uint32(p["threshold"])

        try:
            self._device.modo_standby()
            self._device.modo_configure()

            self._device.enviar(_pack32(20.0),  "10")
            self._device.enviar(_pack32(self._diftemp),      "11")
            self._device.enviar(_pack32(p["tof"]),   "12")
            self._device.enviar(_pack32(p["freq"]),  "14")
            self._device.enviar(_pack32(p["pulse"]), "15")
            self._device.enviar(_pack32(p["pulse"]), "16")
            self._device.enviar(_pack32(p["gain"]),  "17")
            self._device.enviar(_pack32(self._xi),    "18")
            self._device.enviar(_pack32(self._alpha), "19")
            self._device.enviar(_pack32(self._short_corr), "1A")
            self._device.enviar(_pack32(self._long_corr),  "1B")
            self._device.enviar(_pack32(self._short_temp), "1C")
            self._device.enviar(_pack32(self._long_temp),  "1D")
            self._device.enviar(_pack32(long_algo),        "2C")
            self._device.enviar(_pack32(long_umbral),      "2D")

            self._device.modo_save()
            # Dejamos el equipo listo en modo_single para _acquire_frame()
            self._device.modo_standby()
            self._device.modo_single()
        except Exception as e:
            self.error.emit(f"Error enviando config: {e}")
            self._stop = True

    # .................................................................
    def _acquire_frame(self) -> dict:
        try:
            self._device.enviar_temp()
            self._device.start_measure()
            frame = self._device.lectura()
            return frame
        except Exception as e:
            self.error.emit(f"Error en lectura: {e}")
            self._stop = True
            return {}

# ---------------------------------------------------------------------------
# 2) DJTab
# ---------------------------------------------------------------------------
class DJTab(QtWidgets.QWidget):
    """Reemplazo directo para DJTab con actualización fluida."""

    def __init__(self, com_selector: QtWidgets.QComboBox | None = None,
                 db: Optional[Any] = None, batch_id: Optional[str] = None):
        super().__init__()
        self._worker: DeviceWorker | None = None
        self._device: Device | None = None
        # Inject COM selector from MainWindow
        self.com_selector = com_selector
        self.db = db
        self.batch_id = batch_id
        self.last_frame: dict | None = None
        self._manual_stop = False
        self._build_ui()

    # ----------------------------------------------------------- UI factory
    def _build_ui(self):
        layout = QtWidgets.QFormLayout()
        layout.setFormAlignment(Qt.AlignTop)
        layout.setVerticalSpacing(16)

        # Spinners básicos
        self.freq  = labeled_spinbox(2, 100,  step=2, decimals=0)
        self.gain  = labeled_spinbox(5, 80,   step=5, decimals=0)
        self.pulse = labeled_spinbox(1, 16,   step=1, decimals=0)
        self.tof   = labeled_spinbox(0, MAX_32BIT_INT, step=1, decimals=1)
        self.temp  = labeled_spinbox(-20, 50, step=0.1, decimals=1)
        self.bolt  = labeled_spinbox(0, MAX_32BIT_INT, step=1, decimals=0)

        self.freq.setValue(40)
        self.gain.setValue(20)
        self.pulse.setValue(2)
        self.tof.setValue(244500.0)
        self.temp.setValue(20.0)

        # ToF mode widgets
        self.tof_mode = QtWidgets.QComboBox()
        self.tof_mode.addItems(["Manual", "Calculate"])
        # default ToF mode to Calculate
        self.tof_mode.setCurrentIndex(1)
        self.bolt_length_label = QLabel("Bolt Length (mm)")
        self.bolt_length = labeled_spinbox(0, 10000, step=1, decimals=1)
        # self.bolt_length.setValue(720.0)
        self.bolt_length.setValue(714.0)
        self.velocity_label = QLabel("Velocity (m/s)")
        self.velocity = labeled_spinbox(1000, 10000, step=10, decimals=1)
        self.velocity.setValue(5900.0)
        self.bolt_length_label.hide(); self.bolt_length.hide()
        self.velocity_label.hide();    self.velocity.hide()

        # Controles superiores remaquetados en hasta 3 filas (grid)

        # Algoritmo + Threshold
        self.algo_selector = QtWidgets.QComboBox()
        self.algo_selector.addItems(["Absolute maximum", "First maximum", "Second maximum"])
        self.algo_selector.currentIndexChanged.connect(self._on_algo_changed)
        # Nota: 'algo_selector' y 'threshold' se incluyen en el grid inferior

        self.threshold_label = QtWidgets.QLabel("Threshold")
        self.threshold_spin  = QtWidgets.QDoubleSpinBox()
        self.threshold_spin.setRange(0.0, 1e6)
        self.threshold_spin.setDecimals(2)
        self.threshold_spin.setSingleStep(1.0)
        self.threshold_label.setVisible(False)
        self.threshold_spin.setVisible(False)
        # Construir grid con los 11 selectores en 2-3 filas
        top_grid = QtWidgets.QGridLayout()
        top_grid.setHorizontalSpacing(12)
        top_grid.setVerticalSpacing(6)
        pairs = [
            (QtWidgets.QLabel("Frequency"),   self.freq),
            (QtWidgets.QLabel("Gain"),        self.gain),
            (QtWidgets.QLabel("Pulse"),       self.pulse),
            (QtWidgets.QLabel("Temperature"), self.temp),
            (QtWidgets.QLabel("ToF mode"),    self.tof_mode),
            (QtWidgets.QLabel("ToF"),         self.tof),
            (self.bolt_length_label,          self.bolt_length),
            (self.velocity_label,             self.velocity),
            (QtWidgets.QLabel("Bolt Num"),    self.bolt),
            (QtWidgets.QLabel("Algorithm"),   self.algo_selector),
            (self.threshold_label,            self.threshold_spin),
        ]
        # distribuir 6 pares por fila (hasta 2 filas); si faltan, 3 filas
        cols_per_row = 6
        r = 0; c = 0
        for lab, widget in pairs:
            top_grid.addWidget(lab, r, c*2)
            top_grid.addWidget(widget, r, c*2 + 1)
            c += 1
            if c >= cols_per_row:
                c = 0
                r += 1
        layout.addRow(top_grid)

        # Botones
        self.btn_play = QtWidgets.QPushButton("PLAY", objectName="Play")
        self.btn_stop = QtWidgets.QPushButton("STOP", objectName="Stop")
        self.btn_restart_tof = QtWidgets.QPushButton("RESTART TOF GRAPH")
        # Add a distinct color to make it stand out
        self.btn_restart_tof.setStyleSheet(
            "QPushButton{background:#FF9800;color:#FFFFFF;border:none;border-radius:6px;padding:6px 14px;font-weight:600;}"
            "QPushButton:hover{background:#FB8C00;}"
            "QPushButton:pressed{background:#F57C00;}"
            "QPushButton:disabled{background:#BDBDBD;color:#EEEEEE;}"
        )
        self.btn_save = QtWidgets.QPushButton("SAVE", objectName="Save")
        btn_box = QtWidgets.QHBoxLayout();
        btn_box.addWidget(self.btn_play);
        btn_box.addWidget(self.btn_stop);
        btn_box.addWidget(self.btn_restart_tof);
        btn_box.addWidget(self.btn_save)
        layout.addRow(btn_box)

        # Panel derecho (fichero + gráfico)
        param_frame = QtWidgets.QFrame(); param_frame.setLayout(layout)
        right = QtWidgets.QVBoxLayout()

        # Fichero
        file_box = QtWidgets.QHBoxLayout()
        self.file_edit = QtWidgets.QLineEdit()
        # Selector de formato de fichero (CSV o Excel)
        self.format_selector = QComboBox()
        self.format_selector.addItems(["CSV (.csv)", "Excel (.xlsx)"])
        self.file_btn  = QtWidgets.QPushButton("File", objectName="File")
        self.file_btn.clicked.connect(self._choose_file)
        file_box.addWidget(self.file_edit); file_box.addWidget(self.format_selector); file_box.addWidget(self.file_btn)
        right.addLayout(file_box)

        # Métricas
        metric_box = QtWidgets.QHBoxLayout(); metric_box.setSpacing(20)
        metric_box.addWidget(QtWidgets.QLabel("Amp. max:")); self.amp_max = QtWidgets.QLineEdit(); metric_box.addWidget(self.amp_max)
        metric_box.addWidget(QtWidgets.QLabel("Pct. dif:")); self.pct_dif = QtWidgets.QLineEdit(); metric_box.addWidget(self.pct_dif)
        metric_box.addWidget(QtWidgets.QLabel("ToF:"));      self.tof_edit = QtWidgets.QLineEdit(); metric_box.addWidget(self.tof_edit)
        right.addLayout(metric_box)

        # Gráfica
        self.plot = pg.PlotWidget(title="Signal")
        self.plot.setMinimumHeight(330)
        self.plot.showGrid(x=True, y=True, alpha=0.2)
        self.plot.setLabel('bottom', 'Samples'); self.plot.setLabel('left', 'Amplitude')
        self.curve = self.plot.plot([], [], pen=pg.mkPen('#00c853', width=2))
        self.marker = pg.ScatterPlotItem(size=8, pen=pg.mkPen('#ff5722', width=2), brush=pg.mkBrush('#ff5722'), symbol='o')
        self.plot.addItem(self.marker); self.marker.setVisible(False)
        # Disable mouse zoom/pan on main plot to avoid accidental zooms
        try:
            _vb = self.plot.getPlotItem().getViewBox()
            _vb.setMouseEnabled(False, False)
            _vb.setMenuEnabled(False)
            self.plot.wheelEvent = lambda ev: None
            _vb.wheelEvent = lambda ev: None
            _vb.mouseDoubleClickEvent = lambda ev: None
            _vb.mouseDragEvent = lambda ev: None
        except Exception:
            pass
        self.threshold_line = pg.InfiniteLine(angle=0, pen=pg.mkPen('#FFEA00', width=1, style=Qt.DashLine), movable=False)
        self.plot.addItem(self.threshold_line); self.threshold_line.setVisible(False)
        right.addWidget(self.plot, 1)
        # ---- Data Type selector for DJ plot ----
        h_data = QtWidgets.QHBoxLayout()
        h_data.addWidget(QtWidgets.QLabel("Data Type:"))
        self.data_cb = QtWidgets.QComboBox()
        self.data_cb.addItems(["dat2","dat3"])
        self.data_cb.setCurrentText("dat3")
        h_data.addWidget(self.data_cb)
        right.addLayout(h_data)
        # Update plot when data type changes
        self.data_cb.currentIndexChanged.connect(lambda _: self._update_ui(self.last_frame) if self.last_frame else None)

        # Normalize selector (visual only)
        norm_box = QtWidgets.QHBoxLayout()
        norm_box.addWidget(QtWidgets.QLabel("Normalize:"))
        self.normalize_cb = QtWidgets.QComboBox(); self.normalize_cb.addItems(["No", "Yes"])
        self.normalize_cb.setCurrentIndex(0)
        self.normalize_cb.currentIndexChanged.connect(self._on_normalize_changed)
        norm_box.addWidget(self.normalize_cb)
        right.addLayout(norm_box)

        # Axes mode
        axis_box = QtWidgets.QHBoxLayout()
        axis_box.addWidget(QtWidgets.QLabel("Axes mode:"))
        self.axis_mode = QtWidgets.QComboBox(); self.axis_mode.addItems(["Dynamic", "Static"])
        # conecta el cambio de modo de ejes y togglea también visibilidad de Y‐limit
        self.axis_mode.currentIndexChanged.connect(self._on_axis_mode_changed)
        self.axis_mode.currentIndexChanged.connect(lambda i: self.ylim_container.setVisible(i == 1))
        axis_box.addWidget(self.axis_mode)
        right.addLayout(axis_box)

        ybox = QtWidgets.QHBoxLayout(); ybox.addWidget(QtWidgets.QLabel("Y limit:"))
        self.ylim_spin = QtWidgets.QDoubleSpinBox(); self.ylim_spin.setRange(0.0, 1e6); self.ylim_spin.setDecimals(2); self.ylim_spin.setSingleStep(1.0); self.ylim_spin.setAlignment(Qt.AlignCenter); self.ylim_spin.setValue(100000.0)
        self.ylim_spin.valueChanged.connect(self._apply_static_ylim)
        ybox.addWidget(self.ylim_spin)
        self.ylim_container = QtWidgets.QWidget(); self.ylim_container.setLayout(ybox); self.ylim_container.setVisible(False)
        right.addWidget(self.ylim_container)
        # Checkbox para activar plot en vivo de ToF
        tof_chk_layout = QtWidgets.QHBoxLayout()
        self.tof_plot_chk = QtWidgets.QCheckBox("Live ToF plot")
        # Live ToF UI disabled
        # PlotWidget oculto para ToF en vivo
        self.tof_plot_widget = pg.PlotWidget(title="Real-time ToF")
        self.tof_plot_widget.showGrid(x=True, y=True, alpha=0.2)
        self.tof_plot_widget.setLabel('bottom', 'Frame #')
        self.tof_plot_widget.setLabel('left', 'ToF (µs)')
        self.tof_curve = self.tof_plot_widget.plot([], [], pen=pg.mkPen('b', width=1))
        self.tof_plot_widget.setVisible(False)
        # Disable mouse zoom/pan on ToF plot as well
        try:
            _vb2 = self.tof_plot_widget.getPlotItem().getViewBox()
            _vb2.setMouseEnabled(False, False)
            _vb2.setMenuEnabled(False)
            self.tof_plot_widget.wheelEvent = lambda ev: None
            _vb2.wheelEvent = lambda ev: None
            _vb2.mouseDoubleClickEvent = lambda ev: None
            _vb2.mouseDragEvent = lambda ev: None
        except Exception:
            pass
        # Live ToF UI disabled
        # Conectar checkbox para mostrar/ocultar plot ToF
        # Live ToF UI disabled

        # Ajustes del ToF en vivo: color naranja, visible y al lado del gráfico principal
        try:
            self.tof_curve.setPen(pg.mkPen('#FF9800', width=2))
            self.tof_plot_widget.setLabel('left', 'ToF (\u00b5s)')
            self.tof_plot_widget.setVisible(True)
        except Exception:
            pass
        # Recolocar gráficos en una fila
        try:
            right.removeWidget(self.plot)
        except Exception:
            pass
        plots_row = QtWidgets.QHBoxLayout()
        plots_row.addWidget(self.plot, 1)
        plots_row.addWidget(self.tof_plot_widget, 1)
        right.addLayout(plots_row)
        # Reubicar controles inferiores al final: Data, Normalize, Axes, Y limit
        try:
            right.removeItem(h_data)
            right.removeItem(norm_box)
            right.removeItem(axis_box)
            right.removeWidget(self.ylim_container)
        except Exception:
            pass
        bottom_row = QtWidgets.QHBoxLayout()
        bottom_row.addLayout(h_data)
        bottom_row.addLayout(norm_box)
        bottom_row.addLayout(axis_box)
        bottom_row.addWidget(self.ylim_container)
        right.addLayout(bottom_row)
        # Buffers para trazado ToF
        self._tof_x, self._tof_y, self._tof_counter = [], [], 0

        # Layout final
        main = QtWidgets.QVBoxLayout(self)
        main.addWidget(param_frame)
        main.addLayout(right, 1)

        # Señales
        self.btn_play.clicked.connect(self._on_play)
        self.btn_stop.clicked.connect(self._on_stop)
        self.btn_save.clicked.connect(self._on_save)
        # Restart ToF graph button
        self.btn_restart_tof.clicked.connect(self._on_restart_tof)

        # Estado inicial de botones
        self.btn_stop.setEnabled(False)

        # ToF mode logic: show/hide fields and calculate round-trip ToF in µs
        def update_tof():
            # calculate ToF = 2*length(m)/vel(m/s) in µs
            length_mm = float(self.bolt_length.value())
            vel_m_s   = float(self.velocity.value())
            tof_us    = (2*length_mm) / vel_m_s * 1e6
            self.tof.setValue(tof_us)
        def toggle_tof_mode():
            manual = self.tof_mode.currentText() == "Manual"
            self.tof.setEnabled(manual)
            self.bolt_length_label.setVisible(not manual)
            self.bolt_length.setVisible(not manual)
            self.velocity_label.setVisible(not manual)
            self.velocity.setVisible(not manual)
            if not manual:
                update_tof()
        self.tof_mode.currentIndexChanged.connect(lambda _: toggle_tof_mode())
        self.bolt_length.valueChanged.connect(lambda _: update_tof() if self.tof_mode.currentText() == "Calculate" else None)
        self.velocity.valueChanged.connect(lambda _: update_tof() if self.tof_mode.currentText() == "Calculate" else None)
        toggle_tof_mode()

    def _on_restart_tof(self):
        """Reset buffers and clear the live ToF plot."""
        try:
            self._tof_x, self._tof_y, self._tof_counter = [], [], 0
            if hasattr(self, 'tof_curve'):
                self.tof_curve.setData([], [])
        except Exception:
            pass

    # ---------------------------------------------------------------- utils
    def set_com_selector(self, combo: QtWidgets.QComboBox):
        self.com_selector = combo

    def _on_algo_changed(self, idx: int):
        show = idx in (1, 2)
        self.threshold_label.setVisible(show)
        self.threshold_spin.setVisible(show)

    def _on_axis_mode_changed(self, idx: int):
        # muestra/oculta contenedor de Y‐limit
        self.ylim_container.setVisible(idx == 1)
        pi = self.plot.getPlotItem()
        if idx == 1:   # Static
            self.ylim_spin.setEnabled(True)
            y = float(self.ylim_spin.value())
            pi.enableAutoRange(False, False)
            pi.setYRange(-y, y, padding=0)
        else:
            self.ylim_spin.setEnabled(False)
            pi.enableAutoRange(True, True)

    def _apply_static_ylim(self):
        if self.axis_mode.currentIndex() == 1:
            y = float(self.ylim_spin.value())
            self.plot.getPlotItem().setYRange(-y, y, padding=0)

    def _on_normalize_changed(self, idx: int):
        """Cuando se activa la normalización visual, deshabilita Axes mode
        y aplica auto-range; al desactivarla, restaura el comportamiento actual.
        """
        is_yes = (idx == 1)
        try:
            # Deshabilita/rehabilita el selector de ejes
            if hasattr(self, 'axis_mode'):
                self.axis_mode.setEnabled(not is_yes)
            # Oculta el límite Y cuando normaliza
            if hasattr(self, 'ylim_container'):
                self.ylim_container.setVisible(False if is_yes else (self.axis_mode.currentIndex() == 1))
            # Fuerza auto-range cuando normaliza
            pi = self.plot.getPlotItem()
            if is_yes:
                pi.enableAutoRange(True, True)
                # Asegura que el modo de ejes quede en dinámico
                if hasattr(self, 'axis_mode') and self.axis_mode.currentIndex() != 0:
                    self.axis_mode.setCurrentIndex(0)
            else:
                # Restaura según selección actual de Axes mode
                if hasattr(self, 'axis_mode'):
                    self._on_axis_mode_changed(self.axis_mode.currentIndex())
        except Exception:
            pass
        # Refresca la gráfica con/ sin normalización
        if self.last_frame:
            self._update_ui(self.last_frame)

    # ---------------------------------------------------------------- file
    def _choose_file(self):
        # Defino y creo la carpeta "Medidas" un nivel arriba del script
        medidas_dir = Path(__file__).parent.parent / "Medidas"
        medidas_dir.mkdir(parents=True, exist_ok=True)
        # Nombre por defecto dentro de esa carpeta según formato
        fmt = self.format_selector.currentText()
        if fmt.startswith("CSV"):
            default_path = str(medidas_dir / "M64x700.csv")
            filter_str = "CSV files (*.csv)"
        else:
            default_path = str(medidas_dir / "M64x700.xlsx")
            filter_str = "Archivos de Excel (*.xlsx)"
        name, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar como",
            default_path,
            filter_str
        )
        if name and not name.lower().endswith(
            (".xlsx" if filter_str.endswith("*.xlsx)") else ".csv")):
            name += ".xlsx" if filter_str.endswith("*.xlsx)\"") else ".csv"
        self.file_edit.setText(name)

    # ---------------------------------------------------------------- play/stop
    def _on_play(self):
        if not self.com_selector or not self.com_selector.currentText():
            QtWidgets.QMessageBox.critical(self, "COM", "Selecciona un puerto COM válido.")
            return

        try:
            self._device = Device(self.com_selector.currentText(), baudrate=115200, timeout=1)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "UART", f"No se pudo abrir el puerto:\n{e}")
            return

        self.btn_play.setEnabled(False)
        self.btn_stop.setEnabled(True)
        # Live ToF plot eliminado

        self._worker = DeviceWorker(self._device, self._collect_params,
                                   db=self.db, batch_id=self.batch_id, parent=self)
        self._worker.data_ready.connect(self._update_ui, QtCore.Qt.QueuedConnection)
        self._worker.error.connect(self._on_worker_error, QtCore.Qt.QueuedConnection)
        self._worker.start()

    def _on_stop(self):
        self.btn_stop.setEnabled(False)
        self.btn_play.setEnabled(True)
        if self._worker and self._worker.isRunning():
            self._worker.stop()
            self._worker.wait()
            self._worker = None

    # .................................................................
    # DJTab
    def _collect_params(self) -> Dict[str, Any]:
        # Gather parameters for DJTab and log them for debugging
        params = {
            "freq":      float(self.freq.value()),
            "gain":      float(self.gain.value()),
            "pulse":     float(self.pulse.value()),
            "tof":       float(self.tof.value()),
            "temp":      float(self.temp.value()),
            "algo":      self.algo_selector.currentIndex(),
            "threshold": float(self.threshold_spin.value()),
        }
        return params

    # .................................................................
    def _update_ui(self, frame: dict):
        # Puede llegar un dict vacío si hubo error en el hilo
        if not frame:
            return
        self.last_frame = frame

        # -- Métricas --
        self.amp_max.setText(str(frame["maxcorry"]))
        self.pct_dif.setText(f"{frame['porcentaje_diferencia']:.2f}")
        self.tof_edit.setText(f"{frame['tof']:.1f}")

        # -- Gráfica -- (sólo si la pestaña es visible)
        if self.isVisible():
            # Choose correct data array
            dtype = self.data_cb.currentText() if hasattr(self, 'data_cb') else 'dat3'
            data = frame.get(dtype, frame.get('dat3', []))

            # Visual normalization to 1,000,000 if enabled
            plot_data = data
            scale = 1.0
            try:
                if hasattr(self, 'normalize_cb') and self.normalize_cb.currentIndex() == 1:
                    arr = np.asarray(data, dtype=float)
                    maxabs = float(np.max(np.abs(arr))) if arr.size > 0 else 0.0
                    if maxabs > 0.0:
                        scale = 1_000_000.0 / maxabs
                        plot_data = (arr * scale).tolist()
            except Exception:
                plot_data, scale = data, 1.0

            self.curve.setData(plot_data, autoDownsample=True)
            # Actualiza ToF en vivo (naranja)
            try:
                self._tof_counter = getattr(self, '_tof_counter', 0) + 1
                if not hasattr(self, '_tof_x'):
                    self._tof_x, self._tof_y = [], []
                self._tof_x.append(self._tof_counter)
                self._tof_y.append(float(frame.get('tof', 0.0)))
                if len(self._tof_x) > 600:
                    self._tof_x = self._tof_x[-600:]
                    self._tof_y = self._tof_y[-600:]
                if hasattr(self, 'tof_curve'):
                    self.tof_curve.setData(self._tof_x, self._tof_y)
            except Exception:
                pass

            # Threshold + marker
            if self.algo_selector.currentIndex() in (1, 2):
                y0 = float(self.threshold_spin.value())
                self.threshold_line.setValue(y0 * scale)
                self.threshold_line.setVisible(True)
            else:
                self.threshold_line.setVisible(False)

            # Show marker only for dat3
            if dtype == 'dat3':
                self.marker.setData([frame["maxcorrx"]], [frame["maxcorry"] * scale])
                self.marker.setVisible(True)
            else:
                self.marker.setVisible(False)

            # Plot ToF en vivo si está activo
            # Live ToF plot deshabilitado

    # .................................................................
    def _on_worker_error(self, msg: str):
        QtWidgets.QMessageBox.critical(self, "Error", msg)
        self._on_stop()

    # ---------------------------------------------------------------- save
    def _on_save(self):
        """Guarda el último frame en un .xlsx, incluyendo temperatura tras ToF."""
        file_path = self.file_edit.text().strip()
        if not file_path:
            QtWidgets.QMessageBox.warning(self, "Sin fichero", "Selecciona .xlsx donde guardar los datos.")
            return
        if self.last_frame is None:
            QtWidgets.QMessageBox.information(self, "Sin datos", "Aún no se ha recibido ningún frame para guardar.")
            return

        try:
            # ---------- establece DataFrame con el último frame ----------
            frame = self.last_frame
            # Build record with UI params and frame data, include temp from UI
            rec = {
                'Freq': float(self.freq.value()),
                'Gain': float(self.gain.value()),
                'Pulse': float(self.pulse.value()),
                'Bolt Num': float(self.bolt.value()),
                'pico1': frame['pico1'],
                'pct_diff': frame['porcentaje_diferencia'],
                'tof': frame['tof'],
                'temp': frame['temp'],
                'force': frame['force'],
                'maxcorrx': frame['maxcorrx'],
                'maxcorry': frame['maxcorry']
            }
            # Add dat3 and dat2 arrays
            rec.update({f'dat3_{i}': v for i, v in enumerate(frame.get('dat3', []))})
            rec.update({f'dat2_{i}': v for i, v in enumerate(frame.get('dat2', []))})
            df = pd.DataFrame([rec])
            # convierte porcentaje a fracción para formato de salida
            df['pct_diff'] /= 100.0
            # Si CSV, guarda sin formato y sale
            if file_path.lower().endswith('.csv'):
                # Convertir valores de Freq, Gain, Pulse y Bolt Num a int para CSV
                df[['Freq','Gain','Pulse','Bolt Num']] = df[['Freq','Gain','Pulse','Bolt Num']].astype(int)
                csv_path = Path(file_path)
                csv_path.parent.mkdir(parents=True, exist_ok=True)
                mode = 'a' if csv_path.exists() and csv_path.stat().st_size > 0 else 'w'
                header = not (mode == 'a')
                df.to_csv(file_path, index=False, mode=mode, header=header)
                QtWidgets.QMessageBox.information(self, "Guardado", f"Datos guardados en CSV:\n{file_path}")
                return

            # ---------- escribe / añade -------------------------------------------------
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            from openpyxl.styles import PatternFill
            green = PatternFill(fill_type='solid', fgColor='C6EFCE')

            file_exists = Path(file_path).exists()

            if file_exists:
                wb  = load_workbook(file_path)
                ws  = wb['Datos'] if 'Datos' in wb.sheetnames else wb.create_sheet('Datos')
                startrow = ws.max_row
                wb.save(file_path)

                with pd.ExcelWriter(
                        file_path, engine='openpyxl', mode='a',
                        if_sheet_exists='overlay') as writer:
                    df.to_excel(writer, index=False, header=False,
                                startrow=startrow, sheet_name='Datos')
            else:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Datos')

            # ---------- formato de las NUEVAS filas ------------------------------------
            wb  = load_workbook(file_path)
            ws  = wb['Datos']
            pct_col   = df.columns.get_loc('pct_diff') + 1
            tof_col   = df.columns.get_loc('tof')      + 1
            temp_col  = df.columns.get_loc('temp')     + 1
            force_col = df.columns.get_loc('force')    + 1
            max_row   = ws.max_row
            max_col   = ws.max_column
            first_new = max_row - len(df) + 1

            for r in range(first_new, max_row + 1):
                ws.cell(row=r, column=pct_col).number_format   = '0.0%'
                ws.cell(row=r, column=tof_col).number_format   = '0.0'
                ws.cell(row=r, column=temp_col).number_format  = '0.0'
                ws.cell(row=r, column=force_col).number_format = '0.0'
                if ws.cell(row=r, column=pct_col).value > 0.20:
                    for c in range(1, max_col + 1):
                        ws.cell(row=r, column=c).fill = green

            wb.save(file_path)
            QtWidgets.QMessageBox.information(self, "OK", f"Datos guardados en:\n{file_path}")

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Excel", f"No se pudo guardar:\n{e}")
